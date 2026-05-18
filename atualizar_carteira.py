"""
atualizar_carteira.py — v2
═══════════════════════════════════════════════════════════════════════════════
Engine de cálculo da Carteira Clean.

Lê EVENTOS, busca preços de mercado, calcula posições, TWR, atribuição,
projeção, e popula as abas geradas do workbook.

ARQUITETURA:
  - Single source of truth: aba EVENTOS é a verdade. O script nunca a edita.
  - Trade-date accounting (GIPS-aligned)
  - Composites separados: Gerida vs FUNCEF
  - Modelo A retroativo (até 16/05/2026) + Modelo B futuro (CAIXA FIC FUNC como caixa)
  - PEPS corrigido com reset em zeramento
  - TWR com sub-períodos (Modified Dietz simplificado)
  - Idempotência: parte sempre do event log; rodar N vezes = mesmo resultado
  - Validação ativa: detecta inconsistências e emite alertas
  - Auditabilidade: log estruturado de cada cálculo

USO:
  python atualizar_carteira.py                # incremental (default)
  python atualizar_carteira.py --full         # reprocessa do zero
  python atualizar_carteira.py --no-api       # usa só HISTORICO_PRECOS manual
  python atualizar_carteira.py --validate     # só roda validações
  python atualizar_carteira.py --verbose      # log detalhado

DEPENDÊNCIAS:
  pip install yfinance pandas openpyxl requests numpy
═══════════════════════════════════════════════════════════════════════════════
"""

import sys
import argparse
import logging
from pathlib import Path
from datetime import date, datetime, timedelta
from collections import defaultdict
from dataclasses import dataclass

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import yfinance as yf
    import requests
    HAS_NETWORK = True
except ImportError:
    HAS_NETWORK = False

# ═══════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════════════
ARQUIVO = Path(__file__).parent / "Carteira_Clean_v2.xlsx"
DATA_INICIO = date(2026, 1, 2)
DATA_CAIXA_TRANSICAO = date(2026, 5, 17)

COTIZADO_PUBLICO = {"Ação BR", "BDR", "BDR de ETF", "ETF BR"}
COTIZADO_PRIVADO = {"Fundo CP", "Fundo de Pensão", "Fundo Indexado", "Tesouro Direto"}
AGREGADO_PRIVADO = {"Letra de Crédito"}

FLUXOS_EXTERNOS = {"APORTE_EXTERNO", "RESGATE_EXTERNO", "CONTRIBUICAO"}
COMPRAS = {"COMPRA", "SALDO_INICIAL", "BONIFICACAO"}
VENDAS = {"VENDA", "VENCIMENTO"}
PROVENTOS = {"DIVIDENDO", "JCP", "RENDIMENTO", "AMORTIZACAO"}

# ═══════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════
def setup_logging(verbose=False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format='%(asctime)s │ %(levelname)-7s │ %(message)s',
        datefmt='%H:%M:%S',
    )
log = logging.getLogger("carteira")

# ═══════════════════════════════════════════════════════════════
# ESTILOS
# ═══════════════════════════════════════════════════════════════
FONT_NAME = 'Arial'
HEADER_FONT = Font(name=FONT_NAME, bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color='1F4E78')
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color='1F4E78')
SECTION_FILL = PatternFill('solid', start_color='D9E1F2')
TEXT_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color='666666')
RED_FONT = Font(name=FONT_NAME, bold=True, size=10, color='C00000')
GREEN_FONT = Font(name=FONT_NAME, bold=True, size=10, color='006400')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')
thin = Side(border_style='thin', color='BFBFBF')
BORDER = Border(top=thin, left=thin, right=thin, bottom=thin)
FMT_MONEY = 'R$ #,##0.00;(R$ #,##0.00);-'
FMT_PRICE = 'R$ #,##0.0000'
FMT_QTY = '#,##0.000000'
FMT_PCT = '0.00%;(0.00%);-'
FMT_DATE = 'dd/mm/yyyy'

# ═══════════════════════════════════════════════════════════════
# LEITURA
# ═══════════════════════════════════════════════════════════════
def carregar_workbook():
    if not ARQUIVO.exists():
        log.error(f"Arquivo não encontrado: {ARQUIVO}")
        sys.exit(1)
    log.info(f"Lendo {ARQUIVO.name}")
    return openpyxl.load_workbook(ARQUIVO)

def ler_cad_ativos(wb):
    ws = wb["CAD_ATIVOS"]
    ativos = {}
    for r in range(5, ws.max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        if not ticker:
            continue
        ativos[ticker.strip()] = {
            "ticker": ticker.strip(),
            "classe": ws.cell(row=r, column=2).value,
            "familia": ws.cell(row=r, column=3).value,
            "setor": ws.cell(row=r, column=4).value,
            "indexador": ws.cell(row=r, column=5).value,
            "benchmark": ws.cell(row=r, column=6).value,
            "liquidez": ws.cell(row=r, column=7).value,
            "risco": ws.cell(row=r, column=8).value,
            "composite": ws.cell(row=r, column=9).value,
            "observacao": ws.cell(row=r, column=10).value,
        }
    log.info(f"  • {len(ativos)} ativos no CAD_ATIVOS")
    return ativos

def ler_eventos(wb):
    ws = wb["EVENTOS"]
    eventos = []
    for r in range(5, ws.max_row + 1):
        data = ws.cell(row=r, column=1).value
        ativo = ws.cell(row=r, column=2).value
        tipo = ws.cell(row=r, column=3).value
        if not data or not ativo or not tipo:
            continue
        if isinstance(data, datetime):
            d = data.date()
        elif isinstance(data, date):
            d = data
        else:
            continue
        eventos.append({
            "linha": r, "data": d,
            "ativo": ativo.strip(), "tipo": tipo.strip(),
            "qtd": ws.cell(row=r, column=4).value,
            "preco": ws.cell(row=r, column=5).value,
            "valor": ws.cell(row=r, column=6).value or 0,
            "obs": ws.cell(row=r, column=7).value or "",
        })
    eventos.sort(key=lambda e: (e["data"], e["ativo"], e["linha"]))
    log.info(f"  • {len(eventos)} eventos no event log")
    return eventos

def ler_historico_precos(wb):
    ws = wb["HISTORICO_PRECOS"]
    precos = defaultdict(dict)
    for r in range(5, ws.max_row + 1):
        data = ws.cell(row=r, column=1).value
        ticker = ws.cell(row=r, column=2).value
        valor = ws.cell(row=r, column=3).value
        if not data or not ticker or valor is None:
            continue
        if isinstance(data, datetime):
            d = data.date()
        elif isinstance(data, date):
            d = data
        else:
            continue
        precos[ticker.strip()][d] = float(valor)
    log.info(f"  • {sum(len(v) for v in precos.values())} pontos manuais em HISTORICO_PRECOS")
    return dict(precos)

# ═══════════════════════════════════════════════════════════════
# APIs
# ═══════════════════════════════════════════════════════════════
def baixar_precos_publicos(tickers, data_ini, data_fim, no_api=False):
    if no_api or not HAS_NETWORK:
        log.warning("Modo --no-api: pulando download de preços públicos")
        return {}
    out = {}
    log.info(f"Baixando preços via yfinance ({len(tickers)} ativos)…")
    for tkr in tickers:
        yf_tkr = tkr + ".SA"
        try:
            df = yf.download(yf_tkr, start=str(data_ini), end=str(data_fim + timedelta(days=1)),
                             progress=False, auto_adjust=True, threads=False)
            if df.empty:
                log.debug(f"  ✗ {tkr}: vazio")
                continue
            close = df["Close"] if "Close" in df else df.iloc[:, 0]
            if hasattr(close, "squeeze"):
                close = close.squeeze()
            out[tkr] = {pd.Timestamp(idx).date(): float(v) for idx, v in close.items() if pd.notna(v)}
            log.debug(f"  ✓ {tkr}: {len(out[tkr])} dias")
        except Exception as e:
            log.warning(f"  ✗ {tkr}: {e}")
    log.info(f"  → {len(out)}/{len(tickers)} ativos com preços")
    return out

def baixar_benchmarks(data_ini, data_fim, no_api=False):
    if no_api or not HAS_NETWORK:
        return {}
    out = {}
    log.info("Baixando benchmarks…")
    for nome, tkr in [("IBOV", "^BVSP"), ("SP500", "^GSPC"), ("USDBRL", "BRL=X")]:
        try:
            df = yf.download(tkr, start=str(data_ini), end=str(data_fim + timedelta(days=1)),
                             progress=False, auto_adjust=True, threads=False)
            if df.empty:
                continue
            close = df["Close"] if "Close" in df else df.iloc[:, 0]
            if hasattr(close, "squeeze"):
                close = close.squeeze()
            out[nome] = {pd.Timestamp(idx).date(): float(v) for idx, v in close.items() if pd.notna(v)}
            log.debug(f"  ✓ {nome}: {len(out[nome])} dias")
        except Exception as e:
            log.warning(f"  ✗ {nome}: {e}")
    for nome, codigo in [("CDI", 12), ("IPCA", 433)]:
        try:
            url = (f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados"
                   f"?formato=json&dataInicial={data_ini.strftime('%d/%m/%Y')}"
                   f"&dataFinal={data_fim.strftime('%d/%m/%Y')}")
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            df = pd.DataFrame(r.json())
            df["data"] = pd.to_datetime(df["data"], format="%d/%m/%Y").dt.date
            df["valor"] = df["valor"].astype(float) / 100
            out[nome] = dict(zip(df["data"], df["valor"]))
            log.debug(f"  ✓ {nome}: {len(out[nome])} pontos")
        except Exception as e:
            log.warning(f"  ✗ {nome}: {e}")
    return out

# ═══════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════════════════
def preco_em(precos_ticker, d, max_lookback_dias=10):
    if not precos_ticker:
        return None
    if d in precos_ticker:
        return precos_ticker[d]
    datas_validas = sorted(dt for dt in precos_ticker.keys() if dt <= d)
    if not datas_validas:
        return None
    ultima = datas_validas[-1]
    if (d - ultima).days > max_lookback_dias:
        return None
    return precos_ticker[ultima]

def bdate_range(start, end):
    return [d.date() for d in pd.bdate_range(start=start, end=end)]

def num_dias_uteis_entre(d1, d2):
    """Conta dias úteis entre d1 (exclusivo) e d2 (inclusivo). Retorna negativo se d2 < d1."""
    if d2 < d1:
        return -num_dias_uteis_entre(d2, d1)
    if d1 == d2:
        return 0
    return max(0, len(pd.bdate_range(start=d1, end=d2)) - 1)

def status_liquidacao(ev, hoje):
    """Retorna LIQUIDADO / PENDENTE_ENTRADA / PENDENTE_SAIDA.
    Lógica: data + 2 dias úteis ≤ hoje → LIQUIDADO; senão pendente.
    Override manual via obs: 'PENDENTE LIQUIDAÇÃO' força pendente, 'LIQUIDADO' força liquidado."""
    obs = str(ev.get("obs") or "")
    dias_uteis_passados = num_dias_uteis_entre(ev["data"], hoje)
    # Override manual
    if "LIQUIDADO" in obs.upper() and "PENDENTE" not in obs.upper():
        return "LIQUIDADO"
    # Inferência por D+2 (2 dias úteis após data do trade)
    if dias_uteis_passados >= 2:
        return "LIQUIDADO"
    # Pendente — distinguir entrada/saída
    if ev["tipo"] == "COMPRA":
        return "PENDENTE_ENTRADA"
    elif ev["tipo"] == "VENDA":
        return "PENDENTE_SAIDA"
    return "LIQUIDADO"

def data_liquidacao(d_trade):
    """Calcula data de liquidação D+2 a partir da data do trade."""
    dias_uteis = pd.bdate_range(start=d_trade, periods=3)  # trade + 2
    return dias_uteis[-1].date()

# ═══════════════════════════════════════════════════════════════
# INFERÊNCIA DE FLUXOS EXTERNOS RETROATIVOS
# ═══════════════════════════════════════════════════════════════
def _data_efetiva_retroativa(ev):
    """Para eventos agregados provisórios (sem extrato detalhado), antecipa
    a data para o início do mês — refletindo que as movimentações
    aconteceram ao longo do mês, não na data arbitrária do agregado.
    Quando o extrato real chegar e for substituído por eventos detalhados,
    este tratamento deixa de ter efeito (porque a obs não terá mais 'AGREGADO').
    """
    obs = str(ev.get("obs") or "")
    if "AGREGADO" in obs:
        # Antecipa para o 1º dia útil do mês do evento
        d = ev["data"]
        primeiro_dia_mes = date(d.year, d.month, 1)
        # Avança para próximo dia útil
        while primeiro_dia_mes.weekday() >= 5:  # sáb=5, dom=6
            primeiro_dia_mes += timedelta(days=1)
        return primeiro_dia_mes
    return ev["data"]

def inferir_fluxos_externos_retroativos(eventos, ativos, data_corte=DATA_CAIXA_TRANSICAO):
    """
    No período retroativo (antes de data_corte), o caixa NÃO era trackeado.
    Esta função mantém um "caixa virtual" da Carteira Gerida e infere
    APORTE_EXTERNO sempre que uma compra/aplicação não tem cobertura.

    Balanço de caixa virtual:
      Entradas:  VENDAs RV + RESGATEs FIC + PROVENTOS RV
      Saídas:    COMPRAs RV + APLICAÇÕES FIC

    Quando saldo ficaria negativo → APORTE_EXTERNO inferido no valor do déficit.

    Eventos com obs="AGREGADO" têm data antecipada para início do mês
    (usado para tratar resgates consolidados provisórios do FIC FUNC).

    Returns: (lista_aportes_inferidos, saldo_residual)
    """
    saldo_virtual = 0.0
    aportes_inferidos = []

    # Pré-processa: aplica data efetiva (antecipa agregados) e reordena
    eventos_ajustados = []
    for ev in eventos:
        ev_copy = dict(ev)
        ev_copy["_data_efetiva"] = _data_efetiva_retroativa(ev)
        eventos_ajustados.append(ev_copy)
    eventos_ajustados.sort(key=lambda e: (e["_data_efetiva"], e["linha"]))

    for ev in eventos_ajustados:
        d_efetiva = ev["_data_efetiva"]
        if d_efetiva >= data_corte:
            continue
        tkr = ev["ativo"]; tipo = ev["tipo"]
        valor = abs(ev["valor"] or 0)
        familia = ativos.get(tkr, {}).get("familia", "")
        composite = ativos.get(tkr, {}).get("composite", "Gerida")

        if composite == "FUNCEF":
            continue
        if tipo == "SALDO_INICIAL":
            continue

        if tipo == "VENDA" and familia in COTIZADO_PUBLICO:
            saldo_virtual += valor
        elif tipo == "VENDA" and tkr == "CAIXA FIC FUNC":
            saldo_virtual += valor
        elif tipo in PROVENTOS and familia in COTIZADO_PUBLICO:
            saldo_virtual += valor
        elif tipo == "RENDIMENTO" and familia in AGREGADO_PRIVADO:
            pass
        elif tipo == "COMPRA" and familia in COTIZADO_PUBLICO:
            if saldo_virtual < valor:
                falta = valor - saldo_virtual
                aportes_inferidos.append({
                    "data": d_efetiva,
                    "valor": falta,
                    "motivo": f"COMPRA {tkr} R$ {valor:,.2f}",
                })
                saldo_virtual = 0
            else:
                saldo_virtual -= valor
        elif tipo == "COMPRA" and tkr == "CAIXA FIC FUNC":
            if saldo_virtual < valor:
                falta = valor - saldo_virtual
                aportes_inferidos.append({
                    "data": d_efetiva,
                    "valor": falta,
                    "motivo": f"APLICAÇÃO {tkr} R$ {valor:,.2f}",
                })
                saldo_virtual = 0
            else:
                saldo_virtual -= valor
        elif tipo == "COMPRA" and familia == "Tesouro Direto":
            if saldo_virtual < valor:
                falta = valor - saldo_virtual
                aportes_inferidos.append({
                    "data": d_efetiva,
                    "valor": falta,
                    "motivo": f"COMPRA {tkr} R$ {valor:,.2f}",
                })
                saldo_virtual = 0
            else:
                saldo_virtual -= valor

    return aportes_inferidos, saldo_virtual

# ═══════════════════════════════════════════════════════════════
# ENGINE DE POSIÇÕES — PEPS
# ═══════════════════════════════════════════════════════════════
@dataclass
class Posicao:
    qtd: float = 0.0
    custo_total: float = 0.0
    @property
    def custo_medio(self):
        return self.custo_total / self.qtd if self.qtd > 1e-9 else 0.0

def calc_posicoes_e_vendas(eventos, ativos=None):
    """PEPS corrigido com reset em zeramento.
    Returns: (posicoes, vendas_rv, vendas_rf, proventos)
    - vendas_rv: vendas de ativos de Renda Variável (vão para RELATORIO_VENDAS)
    - vendas_rf: 'vendas' de cotizados privados (resgates do FIC FUNC) — para auditoria
    """
    posicoes = defaultdict(Posicao)
    vendas_rv = []
    vendas_rf = []
    proventos = defaultdict(float)
    ativos = ativos or {}
    for ev in eventos:
        tkr = ev["ativo"]; tipo = ev["tipo"]
        p = posicoes[tkr]
        if tipo in COMPRAS or tipo == "CONTRIBUICAO":
            qtd = ev["qtd"] or 0
            valor = ev["valor"] or 0
            p.qtd += qtd
            p.custo_total += abs(valor)
        elif tipo in VENDAS:
            qtd = ev["qtd"] or 0
            valor_recebido = ev["valor"] or 0
            if p.qtd > 1e-9:
                cm = p.custo_medio
                custo_vendido = cm * qtd
                pnl = valor_recebido - custo_vendido
                venda = {
                    "data": ev["data"], "ticker": tkr,
                    "qtd_vendida": qtd, "preco_venda": ev["preco"],
                    "custo_medio": cm, "valor_recebido": valor_recebido,
                    "pnl": pnl,
                    "pnl_pct": pnl/custo_vendido*100 if custo_vendido > 0 else 0
                }
                # Classificar: RV vai pro relatório de vendas;
                # RF/Previdência cotizado é resgate (não venda)
                familia = ativos.get(tkr, {}).get("familia", "")
                if familia in COTIZADO_PUBLICO:
                    vendas_rv.append(venda)
                else:
                    vendas_rf.append(venda)
                p.custo_total -= custo_vendido
                p.qtd -= qtd
                if abs(p.qtd) < 1e-6:
                    p.qtd = 0.0; p.custo_total = 0.0
        elif tipo in PROVENTOS:
            proventos[tkr] += ev["valor"] or 0
    return dict(posicoes), vendas_rv, vendas_rf, dict(proventos)

# ═══════════════════════════════════════════════════════════════
# ENGINE DE EVOLUÇÃO DIÁRIA
# ═══════════════════════════════════════════════════════════════
def calc_evolucao_diaria(eventos, ativos, precos_publicos, precos_manuais, data_fim):
    datas = bdate_range(DATA_INICIO, data_fim)
    if not datas:
        return pd.DataFrame()
    eventos_por_data = defaultdict(list)
    for ev in eventos:
        eventos_por_data[ev["data"]].append(ev)
    posicoes = defaultdict(Posicao)
    linhas = []
    for d in datas:
        for ev in eventos_por_data.get(d, []):
            tkr = ev["ativo"]; tipo = ev["tipo"]
            p = posicoes[tkr]
            qtd = ev["qtd"] or 0; valor = ev["valor"] or 0
            if tipo in COMPRAS or tipo == "CONTRIBUICAO":
                p.qtd += qtd; p.custo_total += abs(valor)
            elif tipo in VENDAS and p.qtd > 1e-9:
                cm = p.custo_medio
                p.custo_total -= cm * qtd
                p.qtd -= qtd
                if abs(p.qtd) < 1e-6:
                    p.qtd = 0.0; p.custo_total = 0.0
        valor_gerida = 0.0
        valor_funcef = 0.0
        valor_rv = 0.0
        for tkr, p in posicoes.items():
            info = ativos.get(tkr, {})
            familia = info.get("familia", "")
            composite = info.get("composite", "Gerida")
            preco = None
            if familia in COTIZADO_PUBLICO and tkr in precos_publicos:
                preco = preco_em(precos_publicos[tkr], d)
            if preco is None:
                preco = preco_em(precos_manuais.get(tkr, {}), d, max_lookback_dias=60)
            if familia in AGREGADO_PRIVADO:
                valor_pos = preco if preco else 0
            elif p.qtd > 1e-9:
                valor_pos = p.qtd * preco if preco else p.custo_total
            else:
                continue
            if composite == "FUNCEF":
                valor_funcef += valor_pos
            else:
                valor_gerida += valor_pos
                if familia in COTIZADO_PUBLICO:
                    valor_rv += valor_pos
        linhas.append({
            "data": d,
            "patrimonio_gerida": valor_gerida,
            "patrimonio_funcef": valor_funcef,
            "patrimonio_total": valor_gerida + valor_funcef,
            "patrimonio_rv": valor_rv,
        })
    return pd.DataFrame(linhas)

# ═══════════════════════════════════════════════════════════════
# TWR + BENCHMARKS
# ═══════════════════════════════════════════════════════════════
def calc_twr_e_benchmarks(df_evo, eventos, benchmarks, aportes_inferidos=None, ativos=None):
    if df_evo.empty:
        return df_evo
    aportes_inferidos = aportes_inferidos or []
    ativos = ativos or {}
    fluxos_g = defaultdict(float); fluxos_f = defaultdict(float)
    fluxos_rv = defaultdict(float)
    # Fluxos externos REGISTRADOS no event log
    for ev in eventos:
        if ev["tipo"] in FLUXOS_EXTERNOS:
            valor = abs(ev["valor"] or 0)
            sinal = 1 if ev["tipo"] in {"APORTE_EXTERNO", "CONTRIBUICAO"} else -1
            if ev["ativo"] == "FUNCEF":
                fluxos_f[ev["data"]] += sinal * valor
            else:
                fluxos_g[ev["data"]] += sinal * valor
        # Fluxos do sub-portfolio RV (compras=entrada, vendas=saída)
        familia = ativos.get(ev["ativo"], {}).get("familia", "")
        if familia in COTIZADO_PUBLICO:
            valor = abs(ev["valor"] or 0)
            if ev["tipo"] == "COMPRA":
                fluxos_rv[ev["data"]] += valor   # entrada no sub-portfolio
            elif ev["tipo"] == "VENDA":
                fluxos_rv[ev["data"]] -= valor   # saída do sub-portfolio
            elif ev["tipo"] == "SALDO_INICIAL":
                fluxos_rv[ev["data"]] += abs(ev["valor"] or 0)  # foto inicial
    # Fluxos externos INFERIDOS retroativamente (apenas Gerida total)
    for ap in aportes_inferidos:
        fluxos_g[ap["data"]] += ap["valor"]

    df = df_evo.copy()
    df["fluxo_gerida"] = df["data"].map(lambda d: fluxos_g.get(d, 0))
    df["fluxo_funcef"] = df["data"].map(lambda d: fluxos_f.get(d, 0))
    df["fluxo_total"] = df["fluxo_gerida"] + df["fluxo_funcef"]
    df["fluxo_rv"] = df["data"].map(lambda d: fluxos_rv.get(d, 0))

    def calc_twr_serie(patrimonio, fluxo):
        retornos = [0.0]
        for i in range(1, len(patrimonio)):
            v0 = patrimonio.iloc[i-1]; v1 = patrimonio.iloc[i]; cf = fluxo.iloc[i]
            denom = v0 + cf
            r = (v1 - denom) / denom if denom > 0 else 0
            retornos.append(r)
        twr_cum = [0.0]
        for r in retornos[1:]:
            twr_cum.append((1 + twr_cum[-1]) * (1 + r) - 1)
        return twr_cum

    df["twr_gerida"] = calc_twr_serie(df["patrimonio_gerida"], df["fluxo_gerida"])
    df["twr_total"] = calc_twr_serie(df["patrimonio_total"], df["fluxo_total"])
    df["twr_rv"] = calc_twr_serie(df["patrimonio_rv"], df["fluxo_rv"]) if "patrimonio_rv" in df.columns else 0

    cdi = benchmarks.get("CDI", {}); ipca = benchmarks.get("IPCA", {})
    ibov = benchmarks.get("IBOV", {}); sp500 = benchmarks.get("SP500", {})
    usdbrl = benchmarks.get("USDBRL", {})

    cdi_acum, ipca_acum, ibov_acum, sp_brl_acum = [], [], [], []
    cdi_v, ipca_v = 1.0, 1.0
    ibov_base = sp_base_brl = None
    for d in df["data"]:
        if d in cdi:
            cdi_v *= (1 + cdi[d])
        cdi_acum.append(cdi_v - 1)
        if d in ipca:
            ipca_v *= (1 + ipca[d])
        ipca_acum.append(ipca_v - 1)
        ibov_p = preco_em(ibov, d)
        if ibov_base is None and ibov_p:
            ibov_base = ibov_p
        ibov_acum.append((ibov_p/ibov_base - 1) if ibov_base and ibov_p else 0)
        sp_p = preco_em(sp500, d); usd = preco_em(usdbrl, d)
        if sp_p and usd:
            sp_brl = sp_p * usd
            if sp_base_brl is None:
                sp_base_brl = sp_brl
            sp_brl_acum.append(sp_brl/sp_base_brl - 1)
        else:
            sp_brl_acum.append(sp_brl_acum[-1] if sp_brl_acum else 0)

    df["cdi_acum"] = cdi_acum
    df["ipca_acum"] = ipca_acum
    df["ibov_acum"] = ibov_acum
    df["sp500_brl_acum"] = sp_brl_acum
    return df

# ═══════════════════════════════════════════════════════════════
# VALIDAÇÃO ATIVA
# ═══════════════════════════════════════════════════════════════
def validar(posicoes, eventos, ativos, df_evo):
    alertas = []
    for tkr, p in posicoes.items():
        if p.qtd < -1e-6:
            alertas.append(("ERRO", tkr, f"Posição negativa ({p.qtd:.4f}) — VENDAs > COMPRAs"))
    ativos_eventos = set(e["ativo"] for e in eventos)
    nao_cadastrados = ativos_eventos - set(ativos.keys())
    for tkr in nao_cadastrados:
        alertas.append(("AVISO", tkr, "Aparece em EVENTOS mas não está cadastrado em CAD_ATIVOS"))
    if not df_evo.empty:
        ultimo = df_evo.iloc[-1]
        pat_gerida = ultimo["patrimonio_gerida"]
        if pat_gerida > 0:
            for tkr, p in posicoes.items():
                if p.qtd < 1e-9: continue
                info = ativos.get(tkr, {})
                if info.get("composite") != "Gerida": continue
                pct = p.custo_total / pat_gerida
                if pct > 0.15:
                    alertas.append(("AVISO", tkr, f"Concentração {pct*100:.1f}% > 15% na Carteira Gerida"))
    for ev in eventos:
        if "PENDENTE LIQUIDAÇÃO" in (ev.get("obs") or ""):
            alertas.append(("INFO", ev["ativo"], f"Evento de {ev['data']} pendente liquidação"))
        if "AGREGADO maio" in (ev.get("obs") or ""):
            alertas.append(("INFO", ev["ativo"], "Evento agregado provisório"))
    return alertas

# ═══════════════════════════════════════════════════════════════
# ESCRITA
# ═══════════════════════════════════════════════════════════════
def limpar_aba(ws, linha_inicio):
    """Limpa conteúdo a partir da linha_inicio. Desmescla células primeiro
    para evitar erro 'MergedCell read-only'."""
    # 1. Desmesclar células no range a limpar
    merged_para_remover = [str(r) for r in ws.merged_cells.ranges
                            if r.min_row >= linha_inicio]
    for rng in merged_para_remover:
        ws.unmerge_cells(rng)
    # 2. Limpar valores e fills
    for r in range(linha_inicio, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

def escrever_posicoes(wb, posicoes, ativos, precos_pub, precos_man, hoje, eventos=None):
    ws = wb["POSICOES"]
    # Cabeçalhos novos com coluna Status
    headers = ["Ticker", "Classe", "Qtd Atual", "Preço Médio", "Custo Total",
               "Preço Atual", "Valor Atual", "P&L R$", "P&L %", "Concentração %", "Status"]
    merged_h = [str(r) for r in ws.merged_cells.ranges if r.min_row == 4]
    for rng in merged_h:
        ws.unmerge_cells(rng)
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=name)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER
        c.border = BORDER

    limpar_aba(ws, 5)

    # Calcular qtd pendente por ativo (entrando ou saindo)
    pendente_por_ativo = defaultdict(lambda: {"entrando": 0, "saindo": 0})
    if eventos:
        for ev in eventos:
            st = status_liquidacao(ev, hoje)
            if st == "PENDENTE_ENTRADA":
                pendente_por_ativo[ev["ativo"]]["entrando"] += ev["qtd"] or 0
            elif st == "PENDENTE_SAIDA":
                pendente_por_ativo[ev["ativo"]]["saindo"] += ev["qtd"] or 0

    pat_total = 0
    rows = []
    for tkr in sorted(posicoes.keys()):
        p = posicoes[tkr]
        info = ativos.get(tkr, {})
        familia = info.get("familia", "")
        if p.qtd < 1e-9 and familia not in AGREGADO_PRIVADO:
            continue
        preco_atual = None
        if familia in COTIZADO_PUBLICO:
            preco_atual = preco_em(precos_pub.get(tkr, {}), hoje)
        if preco_atual is None:
            preco_atual = preco_em(precos_man.get(tkr, {}), hoje, max_lookback_dias=60)
        if familia in AGREGADO_PRIVADO:
            valor_atual = preco_atual if preco_atual else p.custo_total
        else:
            valor_atual = p.qtd * preco_atual if preco_atual else p.custo_total
        pnl = valor_atual - p.custo_total
        pnl_pct = pnl/p.custo_total if p.custo_total > 0 else 0
        pat_total += valor_atual
        # Status
        pend = pendente_por_ativo.get(tkr, {"entrando": 0, "saindo": 0})
        if pend["entrando"] > 0 and pend["saindo"] > 0:
            status = f"🟡 +{pend['entrando']:.0f} / 🔴 −{pend['saindo']:.0f}"
        elif pend["entrando"] > 0:
            status = f"🟡 entrando {pend['entrando']:.0f}"
        elif pend["saindo"] > 0:
            status = f"🔴 saindo {pend['saindo']:.0f}"
        else:
            status = "🟢 liquidado"
        rows.append({"ticker": tkr, "classe": info.get("classe", ""),
                    "qtd": p.qtd, "preco_medio": p.custo_medio,
                    "custo_total": p.custo_total, "preco_atual": preco_atual or 0,
                    "valor_atual": valor_atual, "pnl": pnl, "pnl_pct": pnl_pct,
                    "status": status})
    for i, row in enumerate(rows, start=5):
        conc = row["valor_atual"] / pat_total if pat_total > 0 else 0
        ws.cell(row=i, column=1, value=row["ticker"]).font = BOLD_FONT
        ws.cell(row=i, column=2, value=row["classe"]).font = TEXT_FONT
        c = ws.cell(row=i, column=3, value=row["qtd"]); c.number_format = FMT_QTY
        c = ws.cell(row=i, column=4, value=row["preco_medio"]); c.number_format = FMT_PRICE
        c = ws.cell(row=i, column=5, value=row["custo_total"]); c.number_format = FMT_MONEY
        c = ws.cell(row=i, column=6, value=row["preco_atual"]); c.number_format = FMT_PRICE
        c = ws.cell(row=i, column=7, value=row["valor_atual"]); c.number_format = FMT_MONEY
        c = ws.cell(row=i, column=8, value=row["pnl"]); c.number_format = FMT_MONEY
        c.font = GREEN_FONT if row["pnl"] >= 0 else RED_FONT
        c = ws.cell(row=i, column=9, value=row["pnl_pct"]); c.number_format = FMT_PCT
        c.font = GREEN_FONT if row["pnl_pct"] >= 0 else RED_FONT
        c = ws.cell(row=i, column=10, value=conc); c.number_format = FMT_PCT
        ws.cell(row=i, column=11, value=row["status"]).alignment = LEFT
        for col in range(1, 12):
            ws.cell(row=i, column=col).border = BORDER
            if col not in (1, 2, 11):
                ws.cell(row=i, column=col).alignment = RIGHT
    log.info(f"  ✓ POSICOES: {len(rows)} ativos | Patrimônio total: R$ {pat_total:,.2f}")

def escrever_relatorio_vendas(wb, vendas):
    ws = wb["RELATORIO_VENDAS"]
    limpar_aba(ws, 5)
    total_pnl = 0
    for i, v in enumerate(sorted(vendas, key=lambda x: x["data"]), start=5):
        ws.cell(row=i, column=1, value=v["data"]).number_format = FMT_DATE
        ws.cell(row=i, column=2, value=v["ticker"]).font = BOLD_FONT
        ws.cell(row=i, column=3, value=v["qtd_vendida"]).number_format = FMT_QTY
        ws.cell(row=i, column=4, value=v["preco_venda"]).number_format = FMT_PRICE
        ws.cell(row=i, column=5, value=v["custo_medio"]).number_format = FMT_PRICE
        ws.cell(row=i, column=6, value=v["valor_recebido"]).number_format = FMT_MONEY
        c = ws.cell(row=i, column=7, value=v["pnl"]); c.number_format = FMT_MONEY
        c.font = GREEN_FONT if v["pnl"] >= 0 else RED_FONT
        c = ws.cell(row=i, column=8, value=v["pnl_pct"]/100); c.number_format = FMT_PCT
        c.font = GREEN_FONT if v["pnl"] >= 0 else RED_FONT
        for col in range(1, 9):
            ws.cell(row=i, column=col).border = BORDER
        total_pnl += v["pnl"]
    log.info(f"  ✓ RELATORIO_VENDAS: {len(vendas)} vendas | P&L total: R$ {total_pnl:+,.2f}")

def escrever_evolucao(wb, df_evo):
    ws = wb["EVOLUCAO"]
    # Atualizar cabeçalhos para incluir Patrimônio RV e TWR RV
    headers = ["Data", "Patrim. Gerida", "Patrim. FUNCEF", "Patrim. Total",
               "Patrim. RV", "TWR Gerida", "TWR Total", "TWR RV",
               "CDI", "IBOV", "IPCA", "S&P500 BRL"]
    # Desmesclar antes de escrever cabeçalhos
    merged_h = [str(r) for r in ws.merged_cells.ranges if r.min_row == 4]
    for rng in merged_h:
        ws.unmerge_cells(rng)
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=name)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER
        c.border = BORDER
    limpar_aba(ws, 5)
    for i, row in enumerate(df_evo.itertuples(index=False), start=5):
        ws.cell(row=i, column=1, value=row.data).number_format = FMT_DATE
        ws.cell(row=i, column=2, value=row.patrimonio_gerida).number_format = FMT_MONEY
        ws.cell(row=i, column=3, value=row.patrimonio_funcef).number_format = FMT_MONEY
        ws.cell(row=i, column=4, value=row.patrimonio_total).number_format = FMT_MONEY
        ws.cell(row=i, column=5, value=row.patrimonio_rv).number_format = FMT_MONEY
        ws.cell(row=i, column=6, value=row.twr_gerida).number_format = FMT_PCT
        ws.cell(row=i, column=7, value=row.twr_total).number_format = FMT_PCT
        ws.cell(row=i, column=8, value=row.twr_rv).number_format = FMT_PCT
        ws.cell(row=i, column=9, value=row.cdi_acum).number_format = FMT_PCT
        ws.cell(row=i, column=10, value=row.ibov_acum).number_format = FMT_PCT
        ws.cell(row=i, column=11, value=row.ipca_acum).number_format = FMT_PCT
        ws.cell(row=i, column=12, value=row.sp500_brl_acum).number_format = FMT_PCT
        for col in range(1, 13):
            ws.cell(row=i, column=col).border = BORDER
    log.info(f"  ✓ EVOLUCAO: {len(df_evo)} dias úteis (com TWR RV separado)")

def escrever_dashboard(wb, posicoes, df_evo, ativos, alertas):
    ws = wb["DASHBOARD"]
    if not df_evo.empty:
        ultimo = df_evo.iloc[-1]
        pat_total = ultimo["patrimonio_total"]
        twr_gerida = ultimo["twr_gerida"]
        cdi_ytd = ultimo["cdi_acum"]
        excesso = twr_gerida - cdi_ytd
        n = len(df_evo)
        twr_ann = (1 + twr_gerida) ** (252/n) - 1 if n > 0 else 0
        twr_daily = df_evo["twr_gerida"].diff().fillna(0)
        std = twr_daily.std()
        sharpe = (twr_ann - cdi_ytd) / (std * np.sqrt(252)) if std > 0 else 0
        caixa = 0
        p = posicoes.get("CAIXA FIC FUNC")
        if p:
            caixa = p.custo_total
        ws.cell(row=6, column=1, value=pat_total).number_format = FMT_MONEY
        ws.cell(row=6, column=2, value=f"{twr_gerida*100:+.2f}% / {twr_ann*100:+.2f}% a.a.")
        ws.cell(row=6, column=3, value=excesso).number_format = FMT_PCT
        ws.cell(row=6, column=4, value=round(sharpe, 2))
        ws.cell(row=6, column=5, value=caixa).number_format = FMT_MONEY
        for col in range(1, 6):
            ws.cell(row=6, column=col).font = BOLD_FONT
            ws.cell(row=6, column=col).alignment = CENTER
            ws.cell(row=6, column=col).border = BORDER
    ws.cell(row=22, column=1, value=f"Alertas ativos: {len(alertas)}").font = SECTION_FONT
    for i, (nivel, ativo, msg) in enumerate(alertas[:15], start=23):
        ws.cell(row=i, column=1, value=nivel)
        ws.cell(row=i, column=2, value=ativo).font = BOLD_FONT
        ws.cell(row=i, column=3, value=msg).font = NOTE_FONT
    log.info(f"  ✓ DASHBOARD: KPIs + {len(alertas)} alertas")

# ═══════════════════════════════════════════════════════════════
# CARTEIRA_RV — Estação operacional de Renda Variável
# ═══════════════════════════════════════════════════════════════
def escrever_carteira_rv(wb, posicoes, eventos, ativos, precos_pub, precos_man, df_evo, hoje, benchmarks):
    """4 seções: Caixa operacional / Calendário D+2 / Distribuição setorial / Performance RV."""
    if "CARTEIRA_RV" in wb.sheetnames:
        del wb["CARTEIRA_RV"]
    ws = wb.create_sheet("CARTEIRA_RV", index=1)  # 2ª posição após DASHBOARD

    KPI_TITLE_FONT = Font(name=FONT_NAME, bold=True, size=10, color='4B4B4B')
    KPI_VALUE_FONT = Font(name=FONT_NAME, bold=True, size=14, color='1F4E78')
    KPI_VALUE_RED = Font(name=FONT_NAME, bold=True, size=14, color='C00000')
    KPI_VALUE_GREEN = Font(name=FONT_NAME, bold=True, size=14, color='006400')
    medium = Side(border_style='medium', color='1F4E78')
    KPI_BORDER = Border(top=medium, left=medium, right=medium, bottom=medium)
    GREY_FILL = PatternFill('solid', start_color='F2F2F2')
    RED_FILL = PatternFill('solid', start_color='FCE4D6')

    # === TÍTULO ===
    ws['A1'] = "CARTEIRA_RV — Estação operacional de Renda Variável"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')

    # Contagem de pendências
    pendentes = [ev for ev in eventos if status_liquidacao(ev, hoje) != "LIQUIDADO"
                 and ativos.get(ev["ativo"], {}).get("familia", "") in COTIZADO_PUBLICO]
    n_rv = sum(1 for tkr, p in posicoes.items()
               if p.qtd > 1e-9 and ativos.get(tkr, {}).get("familia") in COTIZADO_PUBLICO)
    ws['A2'] = (f"Foto atualizada em {hoje.strftime('%d/%m/%Y')} | "
                f"{n_rv} ativos RV ativos | {len(pendentes)} operações pendentes D+2")
    ws['A2'].font = NOTE_FONT
    ws.merge_cells('A2:H2')

    # === SEÇÃO 1: CAIXA OPERACIONAL ===
    ws['A4'] = "1. CAIXA OPERACIONAL"
    ws['A4'].font = SECTION_FONT
    ws['A4'].fill = SECTION_FILL
    ws.merge_cells('A4:H4')

    # Caixa atual FIC FUNC (qtd × cota atual)
    caixa_atual = 0.0
    p_fic = posicoes.get("CAIXA FIC FUNC")
    if p_fic and p_fic.qtd > 0:
        cota_fic = preco_em(precos_man.get("CAIXA FIC FUNC", {}), hoje, max_lookback_dias=60)
        if cota_fic:
            caixa_atual = p_fic.qtd * cota_fic

    # Liquidações próximas (entrando / saindo) — D+2 nos próximos 5 dias úteis
    entrando_5d = 0.0
    saindo_5d = 0.0
    proximos_5_du = pd.bdate_range(start=hoje, periods=6).date  # hoje + 5 d.u.
    pendentes_calendario = []
    for ev in eventos:
        if ativos.get(ev["ativo"], {}).get("familia", "") not in COTIZADO_PUBLICO:
            continue
        st = status_liquidacao(ev, hoje)
        if st == "LIQUIDADO":
            continue
        d_liq = data_liquidacao(ev["data"])
        if d_liq not in proximos_5_du:
            continue
        valor = abs(ev["valor"] or 0)
        if st == "PENDENTE_ENTRADA":  # COMPRA → saída de caixa
            saindo_5d += valor
            pendentes_calendario.append((d_liq, ev, -valor))
        elif st == "PENDENTE_SAIDA":  # VENDA → entrada de caixa
            entrando_5d += valor
            pendentes_calendario.append((d_liq, ev, +valor))
    saldo_proj = caixa_atual + entrando_5d - saindo_5d

    kpis = [
        ("CAIXA HOJE (FIC FUNC)", caixa_atual, KPI_VALUE_FONT, GREY_FILL),
        ("LIQUIDAÇÕES ENTRANDO 5d", entrando_5d, KPI_VALUE_GREEN, GREY_FILL),
        ("LIQUIDAÇÕES SAINDO 5d", -saindo_5d, KPI_VALUE_RED, GREY_FILL),
        ("SALDO PROJETADO PÓS-D+2",
         saldo_proj,
         KPI_VALUE_RED if saldo_proj < 0 else KPI_VALUE_GREEN,
         RED_FILL if saldo_proj < 0 else GREEN_FILL),
    ]
    for i, (label, valor, font_value, fill) in enumerate(kpis):
        col_start = 1 + i * 2
        col_end = col_start + 1
        cell_label = ws.cell(row=5, column=col_start, value=label)
        cell_label.font = KPI_TITLE_FONT
        cell_label.alignment = CENTER
        cell_label.fill = GREY_FILL
        ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_end)
        cell_value = ws.cell(row=6, column=col_start, value=valor)
        cell_value.number_format = FMT_MONEY
        cell_value.font = font_value
        cell_value.alignment = CENTER
        cell_value.fill = fill
        ws.merge_cells(start_row=6, start_column=col_start, end_row=6, end_column=col_end)
        for r_ in range(5, 7):
            for c_ in range(col_start, col_end + 1):
                ws.cell(row=r_, column=c_).border = KPI_BORDER
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 36

    # Alerta dinâmico
    if saldo_proj < 0:
        msg = (f"⚠️ ATENÇÃO: caixa projetado fica em R$ {saldo_proj:,.2f} após D+2. "
               f"Precisa resgatar pelo menos R$ {abs(saldo_proj):,.2f} do FIC FUNC.")
        ws['A7'].font = Font(name=FONT_NAME, bold=True, size=10, color='C00000')
    elif saldo_proj < caixa_atual * 0.3:  # menos de 30% do caixa atual
        msg = f"ℹ️ Caixa baixo após liquidações: R$ {saldo_proj:,.2f}. Avaliar resgate preventivo."
        ws['A7'].font = Font(name=FONT_NAME, italic=True, size=10, color='B07000')
    else:
        msg = f"✓ Caixa suficiente para honrar liquidações da semana (R$ {saldo_proj:,.2f} disponível)."
        ws['A7'].font = Font(name=FONT_NAME, italic=True, size=10, color='006400')
    ws['A7'] = msg
    ws['A7'].alignment = LEFT
    ws.merge_cells('A7:H7')

    # === SEÇÃO 2: CALENDÁRIO D+2 ===
    ws['A9'] = "2. CALENDÁRIO DE LIQUIDAÇÕES D+2 (próximos 5 dias úteis)"
    ws['A9'].font = SECTION_FONT
    ws['A9'].fill = SECTION_FILL
    ws.merge_cells('A9:H9')

    cal_cols = ["Liquidação", "Trade", "Op", "Ativo", "Qtd", "Valor R$", "Impacto", "Saldo proj."]
    for col, name in enumerate(cal_cols, start=1):
        c = ws.cell(row=10, column=col, value=name)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER
        c.border = BORDER

    # Ordenar por data de liquidação
    pendentes_calendario.sort(key=lambda x: (x[0], x[1]["data"]))
    saldo_running = caixa_atual
    linha = 11
    if pendentes_calendario:
        for d_liq, ev, impacto in pendentes_calendario:
            saldo_running += impacto
            ws.cell(row=linha, column=1, value=d_liq).number_format = FMT_DATE
            ws.cell(row=linha, column=2, value=ev["data"]).number_format = FMT_DATE
            ws.cell(row=linha, column=3, value=ev["tipo"]).alignment = CENTER
            ws.cell(row=linha, column=4, value=ev["ativo"]).font = BOLD_FONT
            ws.cell(row=linha, column=5, value=ev["qtd"]).alignment = RIGHT
            c = ws.cell(row=linha, column=6, value=abs(ev["valor"] or 0)); c.number_format = FMT_MONEY
            c = ws.cell(row=linha, column=7, value=impacto); c.number_format = FMT_MONEY
            c.font = GREEN_FONT if impacto > 0 else RED_FONT
            c = ws.cell(row=linha, column=8, value=saldo_running); c.number_format = FMT_MONEY
            c.font = RED_FONT if saldo_running < 0 else BOLD_FONT
            for cc in range(1, 9):
                ws.cell(row=linha, column=cc).border = BORDER
            linha += 1
    else:
        ws.cell(row=11, column=1, value="(nenhuma operação pendente nos próximos 5 dias úteis)").font = NOTE_FONT
        ws.merge_cells('A11:H11')
        linha = 12

    # === SEÇÃO 3: DISTRIBUIÇÃO SETORIAL ===
    linha += 2
    ws.cell(row=linha, column=1, value="3. DISTRIBUIÇÃO SETORIAL — Carteira RV (granular temática)").font = SECTION_FONT
    ws.cell(row=linha, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)
    linha += 1

    sec_cols = ["Setor", "Valor R$", "% RV", "# Ativos", "Ativos"]
    for col, name in enumerate(sec_cols, start=1):
        c = ws.cell(row=linha, column=col, value=name)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER
        c.border = BORDER
    ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=8)
    linha += 1

    # Agrupar por setor (valor de mercado, fallback custo)
    setor_valores = defaultdict(lambda: {"valor": 0, "n": 0, "ativos": []})
    total_rv = 0
    for tkr, p in posicoes.items():
        info = ativos.get(tkr, {})
        if info.get("familia") not in COTIZADO_PUBLICO or p.qtd < 1e-9:
            continue
        preco = preco_em(precos_pub.get(tkr, {}), hoje)
        if preco is None:
            preco = preco_em(precos_man.get(tkr, {}), hoje, max_lookback_dias=60)
        valor = p.qtd * preco if preco else p.custo_total
        setor = info.get("setor", "Outros")
        setor_valores[setor]["valor"] += valor
        setor_valores[setor]["n"] += 1
        setor_valores[setor]["ativos"].append(tkr)
        total_rv += valor

    setores_ord = sorted(setor_valores.items(), key=lambda x: -x[1]["valor"])
    for setor, info in setores_ord:
        ws.cell(row=linha, column=1, value=setor).font = BOLD_FONT
        c = ws.cell(row=linha, column=2, value=info["valor"]); c.number_format = FMT_MONEY
        c = ws.cell(row=linha, column=3, value=info["valor"] / total_rv if total_rv > 0 else 0)
        c.number_format = FMT_PCT
        ws.cell(row=linha, column=4, value=info["n"]).alignment = CENTER
        ws.cell(row=linha, column=5, value=", ".join(info["ativos"])).font = NOTE_FONT
        ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=8)
        for cc in range(1, 9):
            ws.cell(row=linha, column=cc).border = BORDER
            if cc == 1: ws.cell(row=linha, column=cc).alignment = LEFT
            elif cc in (2, 3): ws.cell(row=linha, column=cc).alignment = RIGHT
        linha += 1

    # Linha TOTAL
    ws.cell(row=linha, column=1, value="TOTAL RV").font = BOLD_FONT
    ws.cell(row=linha, column=1).fill = GREY_FILL
    c = ws.cell(row=linha, column=2, value=total_rv); c.number_format = FMT_MONEY; c.font = BOLD_FONT
    c.fill = GREY_FILL
    c = ws.cell(row=linha, column=3, value=1.0); c.number_format = FMT_PCT; c.font = BOLD_FONT
    c.fill = GREY_FILL
    c = ws.cell(row=linha, column=4, value=sum(s[1]["n"] for s in setores_ord))
    c.alignment = CENTER; c.font = BOLD_FONT; c.fill = GREY_FILL
    ws.cell(row=linha, column=5).fill = GREY_FILL
    ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=8)
    for cc in range(1, 9):
        ws.cell(row=linha, column=cc).border = BORDER
    linha += 2

    # === SEÇÃO 4: PERFORMANCE RV ===
    ws.cell(row=linha, column=1, value="4. PERFORMANCE DA RV").font = SECTION_FONT
    ws.cell(row=linha, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)
    linha += 1

    ws.cell(row=linha, column=1, value="4.1 Retorno acumulado (YTD)").font = BOLD_FONT
    linha += 1

    # TWR RV + benchmarks (do df_evo)
    if not df_evo.empty:
        ult = df_evo.iloc[-1]
        twr_rv = ult.get("twr_rv", 0)
        ibov = ult.get("ibov_acum", 0)
        sp500 = ult.get("sp500_brl_acum", 0)
        metrics = [
            ("TWR Carteira RV", twr_rv, "Sub-portfolio RV (sem FUNCEF/RF)"),
            ("vs IBOV", ibov, "Benchmark Brasil"),
            ("vs S&P500 BRL", sp500, "Benchmark internacional"),
        ]
        for label, valor, nota in metrics:
            ws.cell(row=linha, column=1, value=label).font = TEXT_FONT
            c = ws.cell(row=linha, column=2, value=valor); c.number_format = FMT_PCT
            c.font = GREEN_FONT if valor >= 0 else RED_FONT
            ws.cell(row=linha, column=3, value=nota).font = NOTE_FONT
            ws.merge_cells(start_row=linha, start_column=3, end_row=linha, end_column=8)
            linha += 1
        linha += 1

        # Top contribuintes do mês (das vendas — proxy)
        ws.cell(row=linha, column=1,
                value=f"4.2 Top contribuintes do mês ({hoje.strftime('%m/%Y')})").font = BOLD_FONT
        linha += 1
        # Vendas realizadas no mês corrente
        mes_atual = (hoje.year, hoje.month)
        # Reusa vendas RV do estado processado (não temos diretamente aqui; ler EVENTOS)
        # Para simplicidade, mostrar contribuição pelos eventos de venda do mês
        ganhos_mes = []
        perdas_mes = []
        pl_por_ativo = defaultdict(float)
        custo_acum = defaultdict(float); qtd_acum = defaultdict(float)
        for ev in sorted(eventos, key=lambda e: e["data"]):
            tkr = ev["ativo"]
            fam = ativos.get(tkr, {}).get("familia", "")
            if fam not in COTIZADO_PUBLICO:
                continue
            if ev["tipo"] in ("COMPRA", "SALDO_INICIAL"):
                qtd_acum[tkr] += ev["qtd"] or 0
                custo_acum[tkr] += abs(ev["valor"] or 0)
            elif ev["tipo"] == "VENDA" and qtd_acum[tkr] > 0:
                cm = custo_acum[tkr] / qtd_acum[tkr]
                pnl = (ev["valor"] or 0) - cm * (ev["qtd"] or 0)
                custo_acum[tkr] -= cm * (ev["qtd"] or 0)
                qtd_acum[tkr] -= ev["qtd"] or 0
                if (ev["data"].year, ev["data"].month) == mes_atual:
                    pl_por_ativo[tkr] += pnl
        positivos = sorted([(t, p) for t, p in pl_por_ativo.items() if p > 0], key=lambda x: -x[1])[:5]
        negativos = sorted([(t, p) for t, p in pl_por_ativo.items() if p < 0], key=lambda x: x[1])[:5]
        ws.cell(row=linha, column=1, value="Positivos:").font = TEXT_FONT
        ws.cell(row=linha, column=1).alignment = LEFT
        if positivos:
            txt = " | ".join(f"{t} +R$ {p:,.2f}" for t, p in positivos)
        else:
            txt = "(nenhum no mês corrente)"
        ws.cell(row=linha, column=2, value=txt).font = GREEN_FONT
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=8)
        linha += 1
        ws.cell(row=linha, column=1, value="Negativos:").font = TEXT_FONT
        if negativos:
            txt = " | ".join(f"{t} R$ {p:,.2f}" for t, p in negativos)
        else:
            txt = "(nenhum no mês corrente)"
        ws.cell(row=linha, column=2, value=txt).font = RED_FONT
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=8)
        linha += 1

    # Larguras
    widths = [26, 16, 14, 10, 16, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    log.info(f"  ✓ CARTEIRA_RV: {n_rv} ativos RV, {len(setores_ord)} setores, "
             f"{len(pendentes_calendario)} liquidações próximas")

# ═══════════════════════════════════════════════════════════════
# ATRIBUIÇÃO MENSAL
# ═══════════════════════════════════════════════════════════════
def calc_atribuicao_mensal(eventos, ativos, precos_pub, precos_man, data_fim):
    """
    Long format: 1 linha por (mês × ativo).
    Para cada mês fechado, calcula:
      - retorno_ativo = (preco_fim_mes / preco_inicio_mes - 1)
      - peso_medio    = média do peso na carteira durante o mês
      - contribuicao  = retorno × peso_medio
    """
    linhas = []
    datas_uteis = bdate_range(DATA_INICIO, data_fim)
    if not datas_uteis:
        return pd.DataFrame()

    # Agrupar dias por mês
    meses = sorted(set((d.year, d.month) for d in datas_uteis))

    # Reconstruir posições + valores para cada dia
    eventos_por_data = defaultdict(list)
    for ev in eventos:
        eventos_por_data[ev["data"]].append(ev)

    posicoes = defaultdict(Posicao)
    valores_diarios = {}  # {(data, ticker): valor}

    def get_preco(tkr, d):
        info = ativos.get(tkr, {})
        familia = info.get("familia", "")
        preco = None
        if familia in COTIZADO_PUBLICO and tkr in precos_pub:
            preco = preco_em(precos_pub[tkr], d)
        if preco is None:
            preco = preco_em(precos_man.get(tkr, {}), d, max_lookback_dias=60)
        return preco, familia

    for d in datas_uteis:
        # Aplicar eventos do dia
        for ev in eventos_por_data.get(d, []):
            tkr = ev["ativo"]; tipo = ev["tipo"]
            p = posicoes[tkr]
            qtd = ev["qtd"] or 0; valor = ev["valor"] or 0
            if tipo in COMPRAS or tipo == "CONTRIBUICAO":
                p.qtd += qtd; p.custo_total += abs(valor)
            elif tipo in VENDAS and p.qtd > 1e-9:
                cm = p.custo_medio
                p.custo_total -= cm * qtd
                p.qtd -= qtd
                if abs(p.qtd) < 1e-6:
                    p.qtd = 0.0; p.custo_total = 0.0
        # Snapshot de valores
        for tkr, p in posicoes.items():
            preco, familia = get_preco(tkr, d)
            if familia in AGREGADO_PRIVADO:
                valor_pos = preco if preco else 0
            elif p.qtd > 1e-9:
                valor_pos = p.qtd * preco if preco else p.custo_total
            else:
                continue
            valores_diarios[(d, tkr)] = valor_pos

    # Para cada mês, agregar
    for ano, mes in meses:
        dias_mes = [d for d in datas_uteis if d.year == ano and d.month == mes]
        if not dias_mes:
            continue
        d_ini = dias_mes[0]
        d_fim = dias_mes[-1]
        # Patrimônio médio Gerida no mês
        pat_medio_gerida = 0
        pat_medio_funcef = 0
        for d in dias_mes:
            pat_d_g = 0; pat_d_f = 0
            for tkr in set(t for (_, t) in valores_diarios.keys()):
                composite = ativos.get(tkr, {}).get("composite", "Gerida")
                v = valores_diarios.get((d, tkr), 0)
                if composite == "FUNCEF":
                    pat_d_f += v
                else:
                    pat_d_g += v
            pat_medio_gerida += pat_d_g
            pat_medio_funcef += pat_d_f
        pat_medio_gerida /= len(dias_mes)
        pat_medio_funcef /= len(dias_mes)

        # Por ativo
        ativos_mes = set(t for d in dias_mes for (dt, t) in valores_diarios.keys() if dt == d)
        for tkr in sorted(ativos_mes):
            v_ini = valores_diarios.get((d_ini, tkr), 0)
            v_fim = valores_diarios.get((d_fim, tkr), 0)
            if v_ini == 0 and v_fim == 0:
                continue
            preco_ini, _ = get_preco(tkr, d_ini)
            preco_fim, _ = get_preco(tkr, d_fim)
            if preco_ini and preco_fim and preco_ini > 0:
                retorno = preco_fim / preco_ini - 1
            else:
                retorno = 0
            peso_medio = (v_ini + v_fim) / 2
            composite = ativos.get(tkr, {}).get("composite", "Gerida")
            denom = pat_medio_gerida if composite == "Gerida" else pat_medio_funcef
            peso_pct = peso_medio / denom if denom > 0 else 0
            contribuicao = retorno * peso_pct
            bench = ativos.get(tkr, {}).get("benchmark", "")
            linhas.append({
                "mes": f"{ano}-{mes:02d}",
                "composite": composite,
                "ativo": tkr,
                "retorno_ativo": retorno,
                "peso_medio": peso_pct,
                "contribuicao": contribuicao,
                "benchmark": bench,
            })

    return pd.DataFrame(linhas)

def escrever_atribuicao_mensal(wb, df_atrib):
    ws = wb["ATRIBUICAO_MENSAL"]
    limpar_aba(ws, 5)
    if df_atrib.empty:
        log.info("  ✓ ATRIBUICAO_MENSAL: sem dados")
        return
    df_atrib = df_atrib.sort_values(["mes", "composite", "ativo"])
    for i, row in enumerate(df_atrib.itertuples(index=False), start=5):
        ws.cell(row=i, column=1, value=row.mes).alignment = CENTER
        ws.cell(row=i, column=2, value=row.composite).alignment = CENTER
        ws.cell(row=i, column=3, value=row.ativo).font = BOLD_FONT
        c = ws.cell(row=i, column=4, value=row.retorno_ativo); c.number_format = FMT_PCT
        c.font = GREEN_FONT if row.retorno_ativo >= 0 else RED_FONT
        c = ws.cell(row=i, column=5, value=row.peso_medio); c.number_format = FMT_PCT
        c = ws.cell(row=i, column=6, value=row.contribuicao); c.number_format = FMT_PCT
        c.font = GREEN_FONT if row.contribuicao >= 0 else RED_FONT
        ws.cell(row=i, column=7, value=row.benchmark).font = NOTE_FONT
        for col in range(1, 8):
            ws.cell(row=i, column=col).border = BORDER
    log.info(f"  ✓ ATRIBUICAO_MENSAL: {len(df_atrib)} linhas (mês × ativo)")

# ═══════════════════════════════════════════════════════════════
# META PATRIMÔNIO
# ═══════════════════════════════════════════════════════════════
def escrever_meta_patrimonio(wb, df_evo, posicoes):
    """Projeção ano-a-ano até atingir R$ 3MM usando TWR janela 12m / YTD anualizada."""
    ws = wb["META_PATRIMONIO"]
    if df_evo.empty:
        return
    # Ler premissas das células amarelas
    meta = ws.cell(row=5, column=2).value or 3000000
    aporte_funcef_mes = ws.cell(row=6, column=2).value or 6392
    aporte_gerida_marco = ws.cell(row=7, column=2).value or 10000
    aporte_gerida_outubro = ws.cell(row=8, column=2).value or 10000

    # Estado inicial
    ultimo = df_evo.iloc[-1]
    pat_atual = ultimo["patrimonio_total"]
    n_dias = len(df_evo)
    twr_ytd = ultimo["twr_total"]
    twr_anualizado = (1 + twr_ytd) ** (252 / n_dias) - 1 if n_dias > 0 else 0
    aporte_anual = 12 * aporte_funcef_mes + aporte_gerida_marco + aporte_gerida_outubro

    # Limpar tabela projeção (linha 14+)
    limpar_aba(ws, 14)

    # Linha 14: cabeçalhos já existem (do build_v2)
    # Linha 15+: projeção
    ano_atual = ultimo["data"].year
    pat_proj = pat_atual
    rows = []
    rows.append((ano_atual, 0, twr_anualizado, pat_proj, pat_proj - meta, pat_proj/meta))
    for ano_offset in range(1, 31):  # até 30 anos
        ano = ano_atual + ano_offset
        pat_proj = pat_proj * (1 + twr_anualizado) + aporte_anual
        gap = pat_proj - meta
        pct_meta = pat_proj / meta
        rows.append((ano, aporte_anual, twr_anualizado, pat_proj, gap, pct_meta))
        if pat_proj >= meta:
            break

    for i, (ano, aporte, twr, pat, gap, pct) in enumerate(rows, start=15):
        ws.cell(row=i, column=1, value=str(ano)).alignment = CENTER
        c = ws.cell(row=i, column=2, value=aporte); c.number_format = FMT_MONEY
        c = ws.cell(row=i, column=3, value=twr); c.number_format = FMT_PCT
        c = ws.cell(row=i, column=4, value=pat); c.number_format = FMT_MONEY
        c.font = BOLD_FONT
        c = ws.cell(row=i, column=5, value=gap); c.number_format = FMT_MONEY
        c.font = GREEN_FONT if gap >= 0 else RED_FONT
        c = ws.cell(row=i, column=6, value=pct); c.number_format = FMT_PCT
        for col in range(1, 7):
            ws.cell(row=i, column=col).border = BORDER
    ano_meta = rows[-1][0] if rows[-1][3] >= meta else "não atinge em 30 anos"
    log.info(f"  ✓ META_PATRIMONIO: meta atingida em {ano_meta} (TWR anualizado: {twr_anualizado*100:+.2f}%)")

# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Engine de cálculo da Carteira Clean")
    parser.add_argument("--full", action="store_true", help="Reprocessa do zero")
    parser.add_argument("--no-api", action="store_true", help="Não chama APIs externas")
    parser.add_argument("--validate", action="store_true", help="Só roda validações")
    parser.add_argument("--verbose", action="store_true", help="Log detalhado")
    args = parser.parse_args()

    setup_logging(args.verbose)
    log.info("═" * 65)
    log.info("Carteira Clean — atualizar_carteira.py v2")
    log.info("═" * 65)

    log.info("\n[1/6] Lendo planilha...")
    wb = carregar_workbook()
    ativos = ler_cad_ativos(wb)
    eventos = ler_eventos(wb)
    precos_manuais = ler_historico_precos(wb)

    log.info("\n[2/6] Baixando preços externos...")
    hoje = date.today()
    tickers_pub = [t for t, info in ativos.items() if info.get("familia") in COTIZADO_PUBLICO]
    precos_publicos = baixar_precos_publicos(tickers_pub, DATA_INICIO, hoje, args.no_api)
    benchmarks = baixar_benchmarks(DATA_INICIO, hoje, args.no_api)

    log.info("\n[3/6] Calculando posições e vendas (PEPS)...")
    posicoes, vendas_rv, vendas_rf, proventos = calc_posicoes_e_vendas(eventos, ativos)
    log.info(f"  • {sum(1 for p in posicoes.values() if p.qtd > 0)} posições ativas")
    log.info(f"  • {len(vendas_rv)} vendas de RV | P&L: R$ {sum(v['pnl'] for v in vendas_rv):+,.2f}")
    log.info(f"  • {len(vendas_rf)} resgates de RF cotizada (auditoria, não entra no rel. vendas)")
    log.info(f"  • R$ {sum(proventos.values()):,.2f} em proventos recebidos")

    log.info("\n[4/6] Reconstruindo evolução diária e TWR...")
    df_evo = calc_evolucao_diaria(eventos, ativos, precos_publicos, precos_manuais, hoje)
    # Inferir fluxos externos retroativos (período antes de 17/05/2026)
    aportes_inferidos, saldo_residual = inferir_fluxos_externos_retroativos(eventos, ativos)
    log.info(f"  • {len(aportes_inferidos)} APORTEs externos inferidos (retroativos)")
    log.info(f"    Total inferido: R$ {sum(a['valor'] for a in aportes_inferidos):,.2f}")
    log.info(f"    Saldo residual em 16/05 (= snapshot caixa Modelo B): R$ {saldo_residual:,.2f}")
    df_evo = calc_twr_e_benchmarks(df_evo, eventos, benchmarks, aportes_inferidos, ativos)
    if not df_evo.empty:
        ult = df_evo.iloc[-1]
        log.info(f"  • Patrimônio em {ult['data']}: R$ {ult['patrimonio_total']:,.2f}")
        log.info(f"    — Gerida: R$ {ult['patrimonio_gerida']:,.2f}")
        log.info(f"    — FUNCEF: R$ {ult['patrimonio_funcef']:,.2f}")
        log.info(f"  • TWR Gerida: {ult['twr_gerida']*100:+.2f}% | CDI: {ult['cdi_acum']*100:+.2f}%")

    log.info("\n[5/6] Calculando atribuição mensal...")
    df_atrib = calc_atribuicao_mensal(eventos, ativos, precos_publicos, precos_manuais, hoje)
    log.info(f"  • {len(df_atrib)} linhas de atribuição (mês × ativo)")

    log.info("\n[6/6] Validações ativas...")
    alertas = validar(posicoes, eventos, ativos, df_evo)
    # Adicionar validação cruzada saldo residual ≈ saldo real
    saldo_real_fic = 2450.48  # reportado pelo user em 16/05/2026
    desvio = abs(saldo_residual - saldo_real_fic)
    if desvio > 100:
        alertas.append(("AVISO", "CAIXA FIC FUNC",
                        f"Saldo residual inferido (R$ {saldo_residual:,.2f}) diverge do reportado (R$ {saldo_real_fic:,.2f}) em R$ {desvio:,.2f}"))
    else:
        alertas.append(("INFO", "CAIXA FIC FUNC",
                        f"✓ Reconciliação caixa virtual ≈ saldo real (desvio R$ {desvio:,.2f})"))
    erros = sum(1 for a in alertas if a[0] == "ERRO")
    avisos = sum(1 for a in alertas if a[0] == "AVISO")
    infos = sum(1 for a in alertas if a[0] == "INFO")
    log.info(f"  • {erros} ERROs | {avisos} AVISOs | {infos} INFOs")
    for nivel, ativo, msg in alertas[:12]:
        log.info(f"    [{nivel}] {ativo}: {msg}")

    if args.validate:
        log.info("\nModo --validate: encerrando sem gravar.")
        return

    log.info("\n[7/7] Gravando abas geradas...")
    escrever_posicoes(wb, posicoes, ativos, precos_publicos, precos_manuais, hoje, eventos)
    escrever_relatorio_vendas(wb, vendas_rv)
    escrever_evolucao(wb, df_evo)
    escrever_atribuicao_mensal(wb, df_atrib)
    escrever_meta_patrimonio(wb, df_evo, posicoes)
    escrever_carteira_rv(wb, posicoes, eventos, ativos, precos_publicos, precos_manuais, df_evo, hoje, benchmarks)
    escrever_dashboard(wb, posicoes, df_evo, ativos, alertas)

    wb.save(ARQUIVO)
    log.info(f"\n✓ Planilha salva: {ARQUIVO.name}")
    log.info("═" * 65)

if __name__ == "__main__":
    main()
