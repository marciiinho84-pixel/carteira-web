"""
Pré-processamento de planilhas (XLSX, CSV) — converte para texto tabular.
"""

import logging

log = logging.getLogger("engine.importacao.planilha")


def xlsx_para_texto(xlsx_bytes: bytes, max_linhas: int = 500) -> str:
    """Converte XLSX para texto tabular legível pelo Claude."""
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        partes = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            partes.append(f"=== Aba: {sheet_name} ===")
            linhas = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_linhas:
                    partes.append(f"[... truncado em {max_linhas} linhas ...]")
                    break
                celulas = [str(c) if c is not None else "" for c in row]
                if any(c for c in celulas):
                    linhas.append("\t".join(celulas))
            partes.append("\n".join(linhas))
        return "\n\n".join(partes)
    except Exception as e:
        log.warning(f"Erro ao ler XLSX: {e}")
        return ""


def csv_para_texto(csv_bytes: bytes, max_linhas: int = 500) -> str:
    """Converte CSV para texto tabular."""
    import csv
    import io
    try:
        # Tenta detectar encoding
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                texto = csv_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return ""
        linhas = texto.splitlines()[:max_linhas]
        return "\n".join(linhas)
    except Exception as e:
        log.warning(f"Erro ao ler CSV: {e}")
        return ""
