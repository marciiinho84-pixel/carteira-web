"""
migrate_from_excel.py — Fase 2.1

Importa as 3 abas-fonte do Carteira_Clean_v2.xlsx para um banco SQLite.
Preserva 100% dos dados; não altera nenhuma lógica de cálculo.

Uso:
    python -m backend.scripts.migrate_from_excel
    python -m backend.scripts.migrate_from_excel --excel /caminho/planilha.xlsx
    python -m backend.scripts.migrate_from_excel --db /caminho/carteira.db
    python -m backend.scripts.migrate_from_excel --dry-run   # valida sem salvar
"""

import sys
import argparse
import logging
from datetime import date, datetime
from pathlib import Path

from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# Garante que o pacote pai seja encontrado independente de onde o script é chamado
ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))

from carteira_clean_web.backend.db.models import Base, Ativo, Evento, PrecoManual

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s │ %(levelname)-7s │ %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("migrate")

TIPOS_VALIDOS = {
    "SALDO_INICIAL", "COMPRA", "VENDA", "DIVIDENDO", "JCP",
    "RENDIMENTO", "AMORTIZACAO", "BONIFICACAO", "CONTRIBUICAO",
    "APORTE_EXTERNO", "RESGATE_EXTERNO", "VENCIMENTO",
}


# ─────────────────────────────────────────────
# Leitura do Excel
# ─────────────────────────────────────────────

def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def ler_cad_ativos(wb):
    ws = wb["CAD_ATIVOS"]
    ativos = []
    erros = []
    for r in range(5, ws.max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        if not ticker:
            continue
        ticker = str(ticker).strip()
        composite = ws.cell(row=r, column=9).value
        if composite not in ("Gerida", "FUNCEF"):
            erros.append(f"  Linha {r}: composite inválido '{composite}' para {ticker}")
            composite = "Gerida"
        ativos.append({
            "ticker": ticker,
            "classe": ws.cell(row=r, column=2).value,
            "familia": ws.cell(row=r, column=3).value,
            "setor": ws.cell(row=r, column=4).value,
            "indexador": ws.cell(row=r, column=5).value,
            "benchmark": ws.cell(row=r, column=6).value,
            "liquidez": ws.cell(row=r, column=7).value,
            "risco": ws.cell(row=r, column=8).value,
            "composite": composite,
            "observacao": ws.cell(row=r, column=10).value,
        })
    return ativos, erros


def ler_eventos(wb, ativos_cadastrados):
    ws = wb["EVENTOS"]
    eventos = []
    erros = []
    for r in range(5, ws.max_row + 1):
        data_v = ws.cell(row=r, column=1).value
        ativo = ws.cell(row=r, column=2).value
        tipo = ws.cell(row=r, column=3).value
        if not data_v or not ativo or not tipo:
            continue
        d = _to_date(data_v)
        if d is None:
            erros.append(f"  Linha {r}: data inválida '{data_v}'")
            continue
        ativo = str(ativo).strip()
        tipo = str(tipo).strip()
        if tipo not in TIPOS_VALIDOS:
            erros.append(f"  Linha {r}: tipo inválido '{tipo}' (ativo={ativo})")
            continue
        if ativo not in ativos_cadastrados:
            erros.append(f"  Linha {r}: ativo '{ativo}' não encontrado em CAD_ATIVOS")
        qtd_v = ws.cell(row=r, column=4).value
        preco_v = ws.cell(row=r, column=5).value
        valor_v = ws.cell(row=r, column=6).value
        obs_v = ws.cell(row=r, column=7).value
        eventos.append({
            "linha_excel": r,
            "data": d,
            "ativo": ativo,
            "tipo": tipo,
            "qtd": float(qtd_v) if qtd_v is not None else None,
            "preco": float(preco_v) if preco_v is not None else None,
            "valor": float(valor_v) if valor_v is not None else 0.0,
            "obs": str(obs_v).strip() if obs_v else "",
        })
    eventos.sort(key=lambda e: (e["data"], e["ativo"], e["linha_excel"]))
    return eventos, erros


def ler_historico_precos(wb):
    ws = wb["HISTORICO_PRECOS"]
    precos = []
    erros = []
    seen = set()
    for r in range(5, ws.max_row + 1):
        data_v = ws.cell(row=r, column=1).value
        ticker = ws.cell(row=r, column=2).value
        valor_v = ws.cell(row=r, column=3).value
        if not data_v or not ticker or valor_v is None:
            continue
        d = _to_date(data_v)
        if d is None:
            erros.append(f"  Linha {r}: data inválida '{data_v}'")
            continue
        ticker = str(ticker).strip()
        chave = (d, ticker)
        if chave in seen:
            erros.append(f"  Linha {r}: duplicata ({ticker}, {d}) — ignorando")
            continue
        seen.add(chave)
        fonte = ws.cell(row=r, column=4).value
        precos.append({
            "data": d,
            "ticker": ticker,
            "valor": float(valor_v),
            "fonte": str(fonte).strip() if fonte else None,
        })
    return precos, erros


# ─────────────────────────────────────────────
# Persistência no SQLite
# ─────────────────────────────────────────────

def criar_banco(db_path: Path):
    engine = create_engine(f"sqlite:///{db_path}", echo=False)
    Base.metadata.create_all(engine)
    return engine


def popular_banco(engine, ativos, eventos, precos, dry_run=False):
    if dry_run:
        log.info("  [dry-run] Nenhuma escrita no banco.")
        return
    with Session(engine) as session:
        # Limpar tabelas antes de repovoar (idempotência)
        session.execute(text("DELETE FROM precos_manuais"))
        session.execute(text("DELETE FROM eventos"))
        session.execute(text("DELETE FROM ativos"))
        session.commit()

        for a in ativos:
            session.add(Ativo(**a))
        session.flush()

        for e in eventos:
            session.add(Evento(**e))
        session.flush()

        for p in precos:
            session.add(PrecoManual(**p))

        session.commit()


# ─────────────────────────────────────────────
# Validação pós-carga
# ─────────────────────────────────────────────

def validar_contagens(engine, esperado_ativos=35, esperado_eventos=105, esperado_precos=47):
    """Verifica se as contagens batem com os valores esperados da Seção 12F do roteiro."""
    with Session(engine) as session:
        n_ativos = session.query(Ativo).count()
        n_eventos = session.query(Evento).count()
        n_precos = session.query(PrecoManual).count()

    ok = True
    resultados = []

    def chk(nome, real, esperado):
        nonlocal ok
        status = "OK" if real == esperado else "DIVERGÊNCIA"
        if real != esperado:
            ok = False
        resultados.append((status, nome, real, esperado))

    chk("Ativos (CAD_ATIVOS)", n_ativos, esperado_ativos)
    chk("Eventos (event log)", n_eventos, esperado_eventos)
    chk("Preços manuais (HISTORICO_PRECOS)", n_precos, esperado_precos)

    return ok, resultados


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Migra Carteira_Clean_v2.xlsx → SQLite")
    parser.add_argument("--excel", default=None, help="Caminho para o .xlsx")
    parser.add_argument("--db", default=None, help="Caminho para o .db de saída")
    parser.add_argument("--dry-run", action="store_true", help="Valida sem gravar no banco")
    args = parser.parse_args()

    # ROOT = /home/marcio/carteira-web  (3 níveis acima de migrate_from_excel.py)
    excel_path = Path(args.excel) if args.excel else ROOT / "Carteira_Clean_v2.xlsx"
    if not excel_path.exists():
        # Fallback: procura na pasta atual
        excel_path = Path("Carteira_Clean_v2.xlsx")
    if not excel_path.exists():
        log.error(f"Planilha não encontrada: {excel_path}")
        sys.exit(1)

    db_path = Path(args.db) if args.db else ROOT / "carteira_clean_web" / "carteira.db"

    log.info("=" * 60)
    log.info("Migração Excel → SQLite — Fase 2.1")
    log.info("=" * 60)
    log.info(f"Fonte : {excel_path}")
    log.info(f"Destino: {db_path}")

    # ── 1. Ler planilha ──────────────────────────────────────────
    log.info("\n[1/4] Lendo planilha Excel...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    ativos, erros_ativos = ler_cad_ativos(wb)
    log.info(f"  • {len(ativos)} ativos lidos de CAD_ATIVOS")
    if erros_ativos:
        log.warning(f"  {len(erros_ativos)} aviso(s) em CAD_ATIVOS:")
        for e in erros_ativos:
            log.warning(e)

    ativos_set = {a["ticker"] for a in ativos}
    eventos, erros_eventos = ler_eventos(wb, ativos_set)
    log.info(f"  • {len(eventos)} eventos lidos de EVENTOS")
    if erros_eventos:
        log.warning(f"  {len(erros_eventos)} aviso(s) em EVENTOS:")
        for e in erros_eventos:
            log.warning(e)

    precos, erros_precos = ler_historico_precos(wb)
    log.info(f"  • {len(precos)} pontos lidos de HISTORICO_PRECOS")
    if erros_precos:
        log.warning(f"  {len(erros_precos)} aviso(s) em HISTORICO_PRECOS:")
        for e in erros_precos:
            log.warning(e)

    wb.close()

    # ── 2. Criar banco ───────────────────────────────────────────
    log.info("\n[2/4] Criando schema SQLite...")
    engine = criar_banco(db_path)
    log.info(f"  ✓ Schema criado: {db_path}")

    # ── 3. Popular banco ─────────────────────────────────────────
    log.info("\n[3/4] Populando banco...")
    popular_banco(engine, ativos, eventos, precos, dry_run=args.dry_run)
    if not args.dry_run:
        log.info(f"  ✓ {len(ativos)} ativos gravados")
        log.info(f"  ✓ {len(eventos)} eventos gravados")
        log.info(f"  ✓ {len(precos)} preços gravados")

    # ── 4. Validar contagens ─────────────────────────────────────
    log.info("\n[4/4] Validando contagens (vs. Seção 12F do roteiro)...")
    if not args.dry_run:
        ok, resultados = validar_contagens(engine)
        for status, nome, real, esperado in resultados:
            simbolo = "✓" if status == "OK" else "✗ DIVERGÊNCIA"
            log.info(f"  {simbolo} {nome}: {real} (esperado {esperado})")
        if ok:
            log.info("\n  ✓ Todas as contagens batem. Migração concluída com sucesso.")
        else:
            log.error("\n  ✗ Há divergências nas contagens. Verificar antes de prosseguir.")
            sys.exit(2)
    else:
        log.info("  [dry-run] Validação de contagens ignorada (sem banco populado).")

    log.info("\n" + "=" * 60)
    log.info("Migração concluída.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
